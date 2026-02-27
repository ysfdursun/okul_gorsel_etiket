"""
Anaokulu Görsel İnceleme Arayüzü — Flask Backend
=================================================
Excel dosyasındaki görsel URL'lerini web arayüzünde gösterir.
Kullanıcı olur/olmaz ile değerlendirir, onaylananlar WebP'ye çevrilip indirilir.

Kullanım:
    pip install flask openpyxl requests pillow
    python reviewer.py

Tarayıcıda: http://localhost:5000
"""

import os
import re
import sys
import io
import time
import requests
from urllib.parse import urlparse
from flask import Flask, render_template, jsonify, request, Response
from openpyxl import load_workbook
from PIL import Image
import threading

# ─── Ayarlar ──────────────────────────────────────────────────────────────────
EXCEL_FILE   = "anaokulu_gorseller.xlsx"
DOWNLOAD_DIR = "indirilen_gorseller"
PORT         = 5000
WEBP_QUALITY = 80   # WebP sıkıştırma kalitesi (1-100)

# Kişi listesi
PERSONS = ["Ahmet", "Emirhan", "Hacer", "Güler", "Fadime"]

app = Flask(__name__)

# ─── Excel Verisi ─────────────────────────────────────────────────────────────

_schools_cache = None

def load_schools():
    """Excel dosyasını okur, okul ve görsel bilgilerini döndürür."""
    global _schools_cache
    if _schools_cache is not None:
        return _schools_cache

    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active

    schools = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
        url    = row[0] if row[0] else ""
        status = row[1] if row[1] else ""
        count  = row[2] if row[2] else 0

        # foto-1, foto-2, ... sütunlarını topla (duplicate kontrolü ile)
        images = []
        seen_urls = set()
        for val in row[3:]:
            if val and str(val).strip():
                img_url = str(val).strip()
                normalized = img_url.split("?")[0].split("#")[0].lower()
                if normalized not in seen_urls:
                    seen_urls.add(normalized)
                    images.append(img_url)

        if not url:
            continue

        domain = urlparse(url).netloc.replace("www.", "")
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', domain)

        schools.append({
            "idx":       row_idx,
            "url":       url,
            "domain":    domain,
            "safe_name": safe_name,
            "status":    status,
            "count":     int(count) if count else 0,
            "images":    images,
        })

    wb.close()
    _schools_cache = schools

    # Kişilere atama: her kişi eşit sayıda GÖRSEL inceleyecek şekilde dağıt
    # Greedy algoritma: en çok görseli olan okuldan başla,
    # her okulu en az görsel yükü olan kişiye ata
    person_load = {p: 0 for p in PERSONS}       # kişi -> toplam görsel sayısı
    person_schools = {p: [] for p in PERSONS}    # kişi -> okul listesi

    # Görseli çok olandan aza doğru sırala
    sorted_schools = sorted(schools, key=lambda s: len(s["images"]), reverse=True)

    for s in sorted_schools:
        # En az yükü olan kişiyi bul
        min_person = min(PERSONS, key=lambda p: person_load[p])
        s["person"] = min_person
        person_load[min_person] += len(s["images"])

    return schools


# ─── İndirme İlerleme Takibi ─────────────────────────────────────────────────

download_progress = {
    "active":    False,
    "school":    "",
    "total":     0,
    "done":      0,
    "errors":    0,
    "finished":  False,
}
progress_lock = threading.Lock()


def convert_to_webp(input_bytes, quality=WEBP_QUALITY):
    """Görsel byte'larını WebP formatına çevirir."""
    try:
        img = Image.open(io.BytesIO(input_bytes))
        # RGBA ise P moduna düşür (WebP destekliyor ama bazı kenarlar sorunlu olabilir)
        if img.mode in ("RGBA", "LA", "PA"):
            pass  # WebP RGBA destekler
        elif img.mode != "RGB":
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="WEBP", quality=quality, method=4)
        return buf.getvalue()
    except Exception:
        return None


def download_images_task(school_name, image_urls, person_name=""):
    """Arka planda görselleri indirir ve WebP'ye çevirir."""
    global download_progress

    # Kişi adı varsa klasörü kişi altında oluştur
    if person_name:
        folder = os.path.join(DOWNLOAD_DIR, person_name, school_name)
    else:
        folder = os.path.join(DOWNLOAD_DIR, school_name)
    os.makedirs(folder, exist_ok=True)

    with progress_lock:
        download_progress = {
            "active":   True,
            "school":   school_name,
            "total":    len(image_urls),
            "done":     0,
            "errors":   0,
            "finished": False,
        }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        "Referer": "",
    }

    for i, img_url in enumerate(image_urls, start=1):
        try:
            h = headers.copy()
            parsed = urlparse(img_url)
            h["Referer"] = f"{parsed.scheme}://{parsed.netloc}/"

            resp = requests.get(img_url, headers=h, timeout=15,
                                verify=False, stream=False)
            resp.raise_for_status()

            # WebP'ye çevir
            webp_data = convert_to_webp(resp.content)
            if webp_data:
                filename = f"foto-{i}.webp"
                filepath = os.path.join(folder, filename)
                with open(filepath, "wb") as f:
                    f.write(webp_data)
            else:
                # Çevrilemezse orijinal olarak kaydet
                path = urlparse(img_url.split("?")[0]).path
                ext = os.path.splitext(path)[1] or ".jpg"
                filename = f"foto-{i}{ext}"
                filepath = os.path.join(folder, filename)
                with open(filepath, "wb") as f:
                    f.write(resp.content)

            with progress_lock:
                download_progress["done"] += 1

        except Exception:
            with progress_lock:
                download_progress["errors"] += 1
                download_progress["done"] += 1

    with progress_lock:
        download_progress["finished"] = True
        download_progress["active"]   = False


# ─── API Rotaları ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/persons")
def api_persons():
    """Kişi listesini ve her kişiye atanan okul/görsel sayısını döndürür."""
    schools = load_schools()
    result = []
    for person in PERSONS:
        p_schools = [s for s in schools if s.get("person") == person]
        p_images = sum(len(s["images"]) for s in p_schools)
        result.append({"name": person, "count": len(p_schools), "images": p_images})
    return jsonify(result)


@app.route("/api/schools")
def api_schools():
    """Okul listesini döndürür. ?person=Ahmet ile filtre."""
    schools = load_schools()
    person_filter = request.args.get("person", "")

    result = []
    for s in schools:
        if person_filter and s.get("person") != person_filter:
            continue
        result.append({
            "idx":       s["idx"],
            "url":       s["url"],
            "domain":    s["domain"],
            "safe_name": s["safe_name"],
            "status":    s["status"],
            "count":     s["count"],
            "person":    s.get("person", ""),
        })
    return jsonify(result)


@app.route("/api/school/<int:idx>/images")
def api_school_images(idx):
    """Bir okulun görsel URL'lerini döndürür."""
    schools = load_schools()
    for s in schools:
        if s["idx"] == idx:
            return jsonify({
                "domain":    s["domain"],
                "safe_name": s["safe_name"],
                "url":       s["url"],
                "images":    s["images"],
                "person":    s.get("person", ""),
            })
    return jsonify({"error": "Okul bulunamadi"}), 404


@app.route("/api/proxy-image")
def api_proxy_image():
    """Görsel URL'sini proxy'ler (CORS sorununu çözer)."""
    img_url = request.args.get("url", "")
    if not img_url:
        return "URL gerekli", 400

    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        }
        resp = requests.get(img_url, headers=headers, timeout=10,
                            verify=False, stream=True)
        content_type = resp.headers.get("Content-Type", "image/jpeg")
        return Response(resp.content, content_type=content_type)
    except Exception:
        return "Gorsel yuklenemedi", 502


@app.route("/api/approve", methods=["POST"])
def api_approve():
    """Onaylanan görselleri arka planda indirir (WebP'ye çevirir)."""
    data = request.get_json()
    school_name = data.get("school_name", "")
    image_urls  = data.get("images", [])
    person_name = data.get("person", "")

    if not school_name or not image_urls:
        return jsonify({"error": "school_name ve images gerekli"}), 400

    if download_progress.get("active"):
        return jsonify({"error": "Zaten bir indirme devam ediyor"}), 409

    t = threading.Thread(target=download_images_task,
                         args=(school_name, image_urls, person_name))
    t.daemon = True
    t.start()

    return jsonify({"message": f"{len(image_urls)} gorsel indiriliyor (WebP)",
                    "school": school_name})


@app.route("/api/progress")
def api_progress():
    """İndirme ilerlemesini döndürür."""
    with progress_lock:
        return jsonify(download_progress.copy())


@app.route("/api/completed")
def api_completed():
    """İndirilmiş okul klasörlerini döndürür (tüm kişi klasörlerini tarar)."""
    completed = set()
    if os.path.exists(DOWNLOAD_DIR):
        for name in os.listdir(DOWNLOAD_DIR):
            path = os.path.join(DOWNLOAD_DIR, name)
            if os.path.isdir(path):
                # Kişi klasörü mü kontrol et
                if name in PERSONS:
                    for subname in os.listdir(path):
                        subpath = os.path.join(path, subname)
                        if os.path.isdir(subpath):
                            files = [f for f in os.listdir(subpath)
                                     if os.path.isfile(os.path.join(subpath, f))]
                            if files:
                                completed.add(subname)
                else:
                    files = [f for f in os.listdir(path)
                             if os.path.isfile(os.path.join(path, f))]
                    if files:
                        completed.add(name)
    return jsonify(list(completed))


# ─── Başlat ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")

    if not os.path.exists(EXCEL_FILE):
        print(f"HATA: {EXCEL_FILE} bulunamadi!")
        sys.exit(1)

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    # Kişi bilgilerini göster
    schools = load_schools()
    print(f"\n{'='*50}")
    print(f"  Gorsel Inceleme Arayuzu")
    print(f"  http://localhost:{PORT}")
    print(f"{'='*50}")
    print(f"\n  Kisi Dagilimi (gorsel bazli):")
    for p in PERSONS:
        p_schools = [s for s in schools if s.get("person") == p]
        p_images = sum(len(s["images"]) for s in p_schools)
        print(f"    {p}: {len(p_schools)} okul, {p_images} gorsel")
    print()

    app.run(debug=False, port=PORT, threaded=True)
