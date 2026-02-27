import hashlib
import time
import re
import logging
import sys
import io
import warnings
from datetime import datetime
from urllib.parse import urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

# ChromeDriver yolunu bir kez indir, sonra cache'den kullan
_chromedriver_path = None

# ─── Ayarlar ──────────────────────────────────────────────────────────────────
MAX_SUBPAGES      = 20       # Her site için max taranacak alt sayfa sayısı
MAX_WORKERS       = 4        # Paralel Selenium thread sayısı (Selenium ağır, 4 yeterli)
PAGE_LOAD_TIMEOUT = 15       # Sayfa yükleme timeout (saniye)
SCROLL_PAUSE      = 1.2      # Lazy-load için scroll arası bekleme (saniye)
MAX_RETRIES       = 2        # Başarısız sayfa için tekrar deneme
URLS_FILE         = "urls.txt"
OUTPUT_FILE       = "anaokulu_gorseller.xlsx"
LOG_FILE          = "scraper.log"

# Aspect ratio limitleri — dışındakiler dekoratif/separator sayılır
MIN_ASPECT_RATIO  = 0.2      # height/width — çok ince yatay bantlar
MAX_ASPECT_RATIO  = 5.0      # height/width — çok uzun dikey şeritler

# Boyut eşiği — her iki boyut da bunun altındaysa atla
MIN_DIMENSION_PX  = 100

# Filtrelenecek domainler
FILTERED_DOMAINS = [
    "instagram.com", "facebook.com", "twitter.com", "youtube.com",
    "tiktok.com", "linkedin.com", "google.com", "business.google.com",
    "sites.google.com", "wixsite.com", "weebly.com", "okul.com.tr",
    "okuloncesigonulluleri.com", "anaokuluvekresler.com", "kursunkalem.com",
    "anaokulu.com.tr", "googlereklamcim.com", "search.google.com", "netlify.app",
]

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg', '.bmp', '.avif'}

# ─── Görsel Filtre Kelimeleri ─────────────────────────────────────────────────

SKIP_URL_KEYWORDS = [
    "logo", "icon", "favicon", "sprite", "badge",
    "btn", "button", "arrow", "close-", "menu-",
    "social", "share", "like", "follow",
    "1x1", "spacer", "pixel", "tracking",
    "thumb-tiny", "banner-small",
    "loading", "spinner", "placeholder",
]

KEEP_URL_KEYWORDS = [
    "galeri", "gallery", "foto", "photo", "resim", "image",
    "slider", "banner", "etkinlik", "activity", "sinif", "class",
    "okul", "school", "upload", "content", "media",
]

GALLERY_PAGE_KEYWORDS = [
    "galeri", "gallery", "fotograf", "foto", "photo",
    "resim", "image", "media", "etkinlik", "activity",
]

# ─── Logging ──────────────────────────────────────────────────────────────────

def setup_logger(log_file):
    lg = logging.getLogger("scraper")
    lg.setLevel(logging.DEBUG)
    if lg.handlers:
        return lg
    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)-7s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    fh = logging.FileHandler(log_file, encoding="utf-8", mode="a")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt)
    lg.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    lg.addHandler(ch)
    return lg

logger = logging.getLogger("scraper")


def log_site_result(url, status, page_count, image_count, skipped_count,
                    dup_count, elapsed_s, error=None):
    domain = get_domain(url)
    if status == "Başarılı":
        logger.info(
            f"[OK]  {domain:<38} | "
            f"sayfa:{page_count:>3}  "
            f"gorsel:{image_count:>4}  "
            f"atlanan:{skipped_count:>4}  "
            f"dup:{dup_count:>3}  "
            f"{elapsed_s:.1f}s"
        )
    else:
        reason = f" — {error}" if error else ""
        logger.warning(f"[--]  {domain:<38} | ULASILAMADI{reason}")


# ─── Yardımcı Fonksiyonlar ────────────────────────────────────────────────────

def is_filtered_url(url):
    domain = urlparse(url.lower()).netloc.replace("www.", "")
    return any(fd in domain for fd in FILTERED_DOMAINS)

def get_domain(url):
    return urlparse(url).netloc.replace("www.", "")

def normalize_url(url):
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "http://" + url
    return url.rstrip("/")

def is_same_domain(url, base_domain):
    try:
        d = urlparse(url).netloc.replace("www.", "")
        return base_domain in d or d in base_domain
    except Exception:
        return False

def is_image_url(url):
    path = urlparse(url.lower().split("?")[0]).path
    return any(path.endswith(ext) for ext in IMAGE_EXTENSIONS)

def url_to_md5(url):
    """URL'yi normalize edip MD5 hash döndürür — duplicate tespiti için."""
    # Query string'i temizle, küçük harfe çevir
    clean = url.lower().split("?")[0].rstrip("/")
    return hashlib.md5(clean.encode("utf-8")).hexdigest()


# ─── Görsel Filtresi ──────────────────────────────────────────────────────────

def should_skip_image(img_url, width=None, height=None):
    """
    True  → atla (ikon, logo, dekoratif, orantısız)
    False → tut
    width/height: piksel cinsinden integer veya None
    """
    url_lower = img_url.lower()
    path = urlparse(url_lower.split("?")[0]).path

    # .ico her zaman atla
    if path.endswith(".ico"):
        return True

    # .svg — ikon kelimesi yoksa tut
    if path.endswith(".svg"):
        has_skip = any(kw in url_lower for kw in SKIP_URL_KEYWORDS)
        has_keep = any(kw in url_lower for kw in KEEP_URL_KEYWORDS)
        return has_skip and not has_keep

    # Galeri/içerik kelimesi varsa kesinlikle tut
    if any(kw in url_lower for kw in KEEP_URL_KEYWORDS):
        return False

    # İkon/logo kelimesi varsa atla
    if any(kw in url_lower for kw in SKIP_URL_KEYWORDS):
        return True

    # Boyut kontrolü
    if width is not None and height is not None:
        # Her ikisi de çok küçük
        if width < MIN_DIMENSION_PX and height < MIN_DIMENSION_PX:
            return True
        # Aspect ratio kontrolü (height / width)
        if width > 0:
            ratio = height / width
            if ratio < MIN_ASPECT_RATIO or ratio > MAX_ASPECT_RATIO:
                return True

    # Tek boyut 32px altı
    if width is not None and width < 32:
        return True
    if height is not None and height < 32:
        return True

    return False


def _parse_dim(val):
    if val is None:
        return None
    try:
        return int(str(val).replace("px", "").strip())
    except (ValueError, AttributeError):
        return None


# ─── Selenium Driver Fabrikası ────────────────────────────────────────────────

def create_driver():
    """Her thread için ayrı headless Chrome driver oluşturur."""
    global _chromedriver_path
    if _chromedriver_path is None:
        _chromedriver_path = ChromeDriverManager().install()

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-images")           # Görselleri yükleme (sadece URL'lerini alıyoruz)
    opts.add_argument("--blink-settings=imagesEnabled=false")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(_chromedriver_path)
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
    )
    return driver


MAX_SCROLLS = 10   # Sonsuz scroll'u önlemek için limit

def scroll_page(driver):
    """
    Sayfayı aşağı kaydırarak lazy-load görsellerini tetikler.
    Max MAX_SCROLLS kez scroll yapar (sonsuz scroll koruması).
    """
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(MAX_SCROLLS):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.3)


# ─── Görsel Çekme (Selenium) ──────────────────────────────────────────────────

def extract_images_selenium(driver, base_url):
    """
    Açık sayfadaki tüm görselleri Selenium + BeautifulSoup ile toplar.
    JS ile render edilmiş görseller ve lazy-load'lar dahildir.
    Returns: (images: set, skipped: int)
    """
    # Sayfayı scroll'layarak lazy-load'ları tetikle
    try:
        scroll_page(driver)
    except Exception:
        pass

    # DOM'u BeautifulSoup'a aktar
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")

    images  = set()
    skipped = 0

    # --- <img> etiketleri ---
    for img in soup.find_all("img"):
        w = _parse_dim(img.get("width"))
        h = _parse_dim(img.get("height"))

        # naturalWidth/naturalHeight Selenium üzerinden de alınabilir ama yavaş
        # HTML attribute yeterliyse kullan
        for attr in ["src", "data-src", "data-lazy-src", "data-original",
                     "data-lazy", "data-srcset", "data-bg"]:
            val = img.get(attr)
            if not val:
                continue
            val = val.split(",")[0].strip().split(" ")[0].strip()
            if not val or val.startswith("data:"):
                continue
            full_url = urljoin(base_url, val)
            if should_skip_image(full_url, w, h):
                skipped += 1
            else:
                images.add(full_url)

    # --- <source> (picture element) ---
    for source in soup.find_all("source"):
        srcset = source.get("srcset") or source.get("data-srcset")
        if not srcset:
            continue
        val = srcset.split(",")[0].strip().split(" ")[0].strip()
        if val and not val.startswith("data:"):
            full_url = urljoin(base_url, val)
            if should_skip_image(full_url):
                skipped += 1
            else:
                images.add(full_url)

    # --- CSS background-image (inline style) ---
    for tag in soup.find_all(style=True):
        urls = re.findall(r'url\(["\']?(.*?)["\']?\)', tag.get("style", ""))
        for u in urls:
            if u and not u.startswith("data:") and is_image_url(u):
                full_url = urljoin(base_url, u)
                if should_skip_image(full_url):
                    skipped += 1
                else:
                    images.add(full_url)

    # --- Selenium: JS ile inject edilmiş background-image'ları da yakala ---
    try:
        elements = driver.find_elements(By.XPATH, "//*[@style]")
        for el in elements:
            style = el.get_attribute("style") or ""
            urls  = re.findall(r'url\(["\']?(.*?)["\']?\)', style)
            for u in urls:
                if u and not u.startswith("data:") and is_image_url(u):
                    full_url = urljoin(base_url, u)
                    if not should_skip_image(full_url):
                        images.add(full_url)
    except Exception:
        pass

    # --- <a> ile direkt görsel linkler ---
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if is_image_url(href):
            full_url = urljoin(base_url, href)
            if should_skip_image(full_url):
                skipped += 1
            else:
                images.add(full_url)

    return images, skipped


def extract_internal_links(soup, base_url, base_domain):
    links = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        full_url = urljoin(base_url, href)
        if is_same_domain(full_url, base_domain):
            links.add(full_url.split("#")[0].rstrip("/"))
    return links


def prioritize_links(links):
    priority, normal = [], []
    for link in links:
        (priority if any(kw in link.lower() for kw in GALLERY_PAGE_KEYWORDS) else normal).append(link)
    return priority + normal


# ─── Hash Duplicate Temizleme ─────────────────────────────────────────────────

def deduplicate_images(image_urls):
    """
    URL tabanlı MD5 hash ile duplicate görselleri temizler.
    Aynı görsel farklı URL'lerden geliyorsa (CDN, boyutlu versiyonlar)
    ilk gördüğünü tutar, diğerlerini atar.
    Returns: (unique_urls: list, dup_count: int)
    """
    seen_hashes = set()
    unique      = []
    dup_count   = 0

    for url in image_urls:
        h = url_to_md5(url)
        if h in seen_hashes:
            dup_count += 1
        else:
            seen_hashes.add(h)
            unique.append(url)

    return unique, dup_count


# ─── Ana Tarama Fonksiyonu ────────────────────────────────────────────────────

def scrape_site(url):
    """
    Bir siteyi Selenium ile tarar.
    Her page için: yükle → scroll → görsel topla → linkleri çıkar
    Site bitince hash duplicate temizle → log → sonuç döndür
    """
    site_start    = time.time()
    base_domain   = get_domain(url)
    all_images    = set()
    total_skipped = 0
    visited       = set()
    to_visit      = [normalize_url(url)]
    page_count    = 0
    driver        = None

    try:
        driver = create_driver()

        while to_visit and page_count < MAX_SUBPAGES:
            current_url = to_visit.pop(0)
            normalized  = normalize_url(current_url)
            if normalized in visited:
                continue
            visited.add(normalized)

            # Sayfayı yükle (retry ile)
            loaded = False
            for attempt in range(MAX_RETRIES):
                try:
                    driver.get(current_url)
                    # Temel DOM yüklenene kadar bekle
                    WebDriverWait(driver, PAGE_LOAD_TIMEOUT).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                    loaded = True
                    break
                except TimeoutException:
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(2)
                except WebDriverException as e:
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(2)

            if not loaded:
                if page_count == 0:
                    # Ana sayfaya bile ulaşılamadı
                    elapsed = time.time() - site_start
                    log_site_result(url, "Ulaşılamadı", 0, 0, 0, 0, elapsed,
                                    "sayfa yuklenemedi")
                    return (url, "Ulaşılamadı", [])
                continue

            page_count += 1

            # Görselleri topla
            images, skipped = extract_images_selenium(driver, current_url)
            all_images.update(images)
            total_skipped += skipped

            # Internal linkleri çıkar (limit aşılmadıysa)
            if page_count < MAX_SUBPAGES:
                try:
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    internal_links = extract_internal_links(soup, current_url, base_domain)
                    new_links = prioritize_links([
                        lnk for lnk in internal_links
                        if normalize_url(lnk) not in visited
                    ])
                    to_visit = new_links + to_visit
                except Exception:
                    pass

    except Exception as e:
        elapsed = time.time() - site_start
        log_site_result(url, "Ulaşılamadı", page_count, 0, 0, 0, elapsed, str(e)[:80])
        return (url, "Ulaşılamadı", [])
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    elapsed = time.time() - site_start

    if page_count == 0:
        log_site_result(url, "Ulaşılamadı", 0, 0, 0, 0, elapsed, "hic sayfa yuklenemedi")
        return (url, "Ulaşılamadı", [])

    # Hash ile duplicate temizle (her site bitince anlık)
    sorted_images          = sorted(all_images)
    unique_images, dup_cnt = deduplicate_images(sorted_images)

    log_site_result(url, "Başarılı", page_count, len(unique_images),
                    total_skipped, dup_cnt, elapsed)

    return (url, "Başarılı", unique_images)


# ─── URL Yükleme ──────────────────────────────────────────────────────────────

def load_urls(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()
    seen, urls = set(), []
    for line in lines:
        url = line.strip()
        if not url:
            continue
        norm = normalize_url(url)
        if is_filtered_url(norm) or norm in seen:
            continue
        seen.add(norm)
        urls.append(url)
    return urls


# ─── Excel Çıktısı ────────────────────────────────────────────────────────────

def create_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Anaokulu Görselleri"

    max_images = max((len(r[2]) for r in results), default=0)
    headers    = ["URL", "Durum", "Görsel Sayısı"] + [f"foto-{i}" for i in range(1, max_images + 1)]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_num, value=header)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border

    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill    = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    for row_num, (url, status, images) in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=url)
        ws.cell(row=row_num, column=2, value=status)
        ws.cell(row=row_num, column=3, value=len(images))
        fill = success_fill if status == "Başarılı" else fail_fill
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).fill   = fill
            ws.cell(row=row_num, column=col).border = thin_border
        for img_idx, img_url in enumerate(images, 1):
            ws.cell(row=row_num, column=3 + img_idx, value=img_url)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    for i in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 60

    ws.freeze_panes    = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(output_path)
    logger.info(f"Excel kaydedildi: {output_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    warnings.filterwarnings("ignore")
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    setup_logger(LOG_FILE)

    test_mode = "--test" in sys.argv
    run_ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    logger.info("=" * 70)
    logger.info(f"  Anaokulu Gorsel Toplama — {run_ts}{'  [TEST]' if test_mode else ''}")
    logger.info(f"  Motor: Selenium (headless Chrome) | Filtreler: boyut + aspect ratio + hash dedup")
    logger.info("=" * 70)

    logger.info(f"URL dosyasi yukleniyor: {URLS_FILE}")
    urls = load_urls(URLS_FILE)
    logger.info(f"Benzersiz URL sayisi: {len(urls)}")

    if test_mode:
        urls = urls[:5]
        logger.info(f"Test modu: sadece ilk {len(urls)} site taranacak")

    logger.info(
        f"Tarama basladi — {MAX_WORKERS} paralel Selenium thread | "
        f"site basina max {MAX_SUBPAGES} sayfa"
    )
    logger.info("-" * 70)
    logger.info(
        f"{'DURUM':<6} {'DOMAIN':<38}  "
        f"{'SAYFA':>5}  {'GORSEL':>6}  {'ATLANAN':>7}  {'DUP':>4}  SURE"
    )
    logger.info("-" * 70)

    results    = []
    completed  = 0
    total      = len(urls)
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_url = {executor.submit(scrape_site, url): url for url in urls}
        for future in as_completed(future_to_url):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                url = future_to_url[future]
                results.append((url, "Ulaşılamadı", []))
                logger.warning(f"[!!]  {get_domain(url)} — beklenmedik hata: {e}")

            completed += 1

            if completed % 10 == 0 or completed == total:
                elapsed    = time.time() - start_time
                rate       = completed / elapsed if elapsed > 0 else 0
                eta        = (total - completed) / rate if rate > 0 else 0
                success    = sum(1 for r in results if r[1] == "Başarılı")
                total_imgs = sum(len(r[2]) for r in results)
                logger.info(
                    f">>> [{completed:>{len(str(total))}}/{total}] "
                    f"basarili:{success:>4}  "
                    f"toplam gorsel:{total_imgs:>5}  "
                    f"gecen:{elapsed:>5.0f}s  "
                    f"kalan:~{eta:>4.0f}s"
                )

    elapsed_total = time.time() - start_time

    url_order = {u: i for i, u in enumerate(urls)}
    results.sort(key=lambda r: url_order.get(r[0], 0))

    success_count = sum(1 for r in results if r[1] == "Başarılı")
    fail_count    = len(results) - success_count
    total_images  = sum(len(r[2]) for r in results)

    logger.info("=" * 70)
    logger.info("  SONUC OZETI")
    logger.info(f"  Basarili site    : {success_count}")
    logger.info(f"  Ulasilamayan     : {fail_count}")
    logger.info(f"  Toplam gorsel    : {total_images}")
    logger.info(f"  Toplam sure      : {elapsed_total:.1f}s  ({elapsed_total / 60:.1f} dakika)")
    logger.info("=" * 70)

    logger.info("Excel dosyasi olusturuluyor...")
    create_excel(results, OUTPUT_FILE)
    logger.info("Tum islemler tamamlandi.")


if __name__ == "__main__":
    main()